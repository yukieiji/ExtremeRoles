import os
import json
from typing import Dict
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

EXTREMERORLS_IN_FILE = os.path.join(WORKING_DIR, "ExtremeRolesTransData.xlsx")
EXTREMERORLS_OUT_FILE = os.path.join(WORKING_DIR, "ExtremeRoles", "Resources", "JsonData", "Language.json")

EXTREMESKIN_IN_FILE = os.path.join(WORKING_DIR, "ExtremeSkinsTransData.xlsx")
EXTREMESKIN_OUT_FILE = os.path.join(WORKING_DIR, "ExtremeSkins", "Resources", "LangData", "stringData.json")

def is_require_update(new_json : Dict[str, str], output_file : str) -> bool:

  try:
    with open(output_file, "r") as f:
      old_json = json.load(f)
    old_json_str = json.dumps(old_json)
  except Exception:
    return True
  else:
    new_json_str = json.dumps(new_json)

    return old_json_str != new_json_str


def xlsx_to_json(file_name : str, output_file : str) -> None:

  wb = load_workbook(file_name, read_only = True)

  xlsx_data = {}

  for s in wb:
    rows = s.iter_rows(min_col=1, min_row=2, max_col=17, max_row=None)
    headers = []
    for header in s[1]:
      if header.value:
        headers.append(header.value)

    for row in rows:
      name = row[0].value

      if not name:
        continue

      data = {}

      for i, cell in enumerate(row[1:]):
        if cell.value:
          # I hate excel why did I do this to myself
          data[str(i)] = cell.value.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")

      if data:
        xlsx_data[name] = data

  if is_require_update(xlsx_data, output_file):

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    with open(output_file, "w") as f:
      json.dump(xlsx_data, f, indent=4)


if __name__ == "__main__":
  xlsx_to_json(EXTREMERORLS_IN_FILE, EXTREMERORLS_OUT_FILE)
  xlsx_to_json(EXTREMESKIN_IN_FILE, EXTREMESKIN_OUT_FILE)