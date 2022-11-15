import os
import subprocess
from typing import Dict, List
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

EXTREMERORLS_IN_FILE = os.path.join(WORKING_DIR, "ExtremeRolesTransData.xlsx")
EXTREMERORLS_OUT_FILE = os.path.join(WORKING_DIR, "ExtremeRoles", "Resources", "JsonData", "Language.json")

EXTREMESKIN_IN_FILE = os.path.join(WORKING_DIR, "ExtremeSkinsTransData.xlsx")
EXTREMESKIN_OUT_FILE = os.path.join(WORKING_DIR, "ExtremeSkins", "Resources", "LangData", "stringData.json")

SUPPORT_LANG = {0:'English', 11:'Japanese', 13: 'SChinese'}
TAG = 'report'

def get_trans_data_check(filename:str)-> Dict[str, List[str]]:
  wb = load_workbook(filename, read_only = True)

  data = {}
  for s in wb:
    rows = s.iter_rows(min_col = 1, min_row = 2, max_col = 17, max_row = None)
    for row in rows:
      name = row[0].value

      if not name:
        continue

      for i, string in enumerate(row[1:]):

        # 日本語の翻訳者は要らないのですきっぽー
        if (not (i in SUPPORT_LANG) or
            (i == 11 and (name == 'langTranslate' or name == 'translatorMember'))):
          continue

        key = SUPPORT_LANG[i]

        if not (key in data):
          data[key] = []

        if not string.value:
          data[key].append(name)

  return data

def convert_md_table(data: Dict[str, List[str]]) -> str:
  result = '| Languages | Missing TransKeys |\n| --- | --- |'

  for lang, miss_keys in data.items():
    if len(miss_keys) == 0:
      result = f'{result}\n| {lang} | ☑ |'
      continue
    for index, key in enumerate(miss_keys):
      line = '\n'
      if index == 0:
        line = f'{line}| {lang} | {key} |'
      else:
        line = f'{line}|>     | {key} |'
      result = f'{result}{line}'

  return result

def output_md_report(*check_xlsx_file):

  result = '## Translation Checker Report\n'

  for file in check_xlsx_file:
    result = f'{result}\n - FileName:{os.path.basename(file)}\n\n'
    result = f'{result}{convert_md_table(get_trans_data_check(file))}\n'

  echo_command = f'echo ("{TAG}={result}" >> $env:GITHUB_ENV'
  subprocess.call(echo_command, shell=True)

if __name__ == "__main__":
  output_md_report(EXTREMERORLS_IN_FILE, EXTREMESKIN_IN_FILE)