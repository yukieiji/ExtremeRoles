import os
import json
import tempfile
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

EXTREMERORLS_IN_FILE = os.path.join(WORKING_DIR, "ExtremeRolesTransData.xlsx")
EXTREMERORLS_OUT_FILE = os.path.join(WORKING_DIR, "ExtremeRoles", "Resources", "JsonData", "Language.json")

EXTREMESKIN_IN_FILE = os.path.join(WORKING_DIR, "ExtremeSkinsTransData.xlsx")
EXTREMESKIN_OUT_FILE = os.path.join(WORKING_DIR, "ExtremeSkins", "Resources", "LangData", "stringData.json")

def writeJsonIfNeed(stringData, outputFile, **kwargs):
  with tempfile.TemporaryFile("w+") as f:
    json.dump(stringData, f, **kwargs)
    f.seek(0)
    currentdata=f.read()
  try:
    with open(outputFile, "r") as f:
      olddata=f.read()
  except Exception:
    need = True
  else:
    need = currentdata!=olddata

  if need:
    os.makedirs(os.path.dirname(outputFile), exist_ok=True)
    with open(outputFile, "w") as f:
      f.write(currentdata)


def stringToJson(filename, outputFile):
  wb = load_workbook(filename, read_only = True)

  stringData = {}
  for s in wb:
    rows = s.iter_rows(min_col = 1, min_row = 2, max_col = 17, max_row = None)
    headers = []
    for header in s[1]:
      if header.value:
        headers.append(header.value)

    for row in rows:
      name = row[0].value

      if not name:
        continue

      data = {}

      for i, string in enumerate(row[1:]):
        if string.value:
          # I hate excel why did I do this to myself
          data[i] = string.value.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")

      if data:
        stringData[name] = data

  writeJsonIfNeed(stringData, outputFile, indent=4)

if __name__ == "__main__":
  stringToJson(EXTREMERORLS_IN_FILE, EXTREMERORLS_OUT_FILE)
  stringToJson(EXTREMESKIN_IN_FILE, EXTREMESKIN_OUT_FILE)